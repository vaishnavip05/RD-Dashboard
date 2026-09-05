"""Generate the browser data file from the latest FLABS research workbooks."""
from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

RESEARCH = Path(r"C:\Users\Dell\Downloads\fwdresearchdataconsolidatedcountason10082026")
YEARWISE = Path(r"C:\Users\Dell\Downloads\fwdupdatedconsolidatedcountyearwisefundedprojecty")
OUT = Path(__file__).parent / "data.js"


def clean(value):
    return "" if value is None else re.sub(r"\s+", " ", str(value)).strip()


def department(value):
    raw = clean(value).upper()
    aliases = {
        "CS": "Computer Science", "COMPUTER SCIENCE": "Computer Science", "DEPARTMENT OF COMPUTER SCIENCE": "Computer Science",
        "MATHS": "Mathematics", "MATHEMATICS": "Mathematics", "BIO TECH": "Biotechnology", "BIO TECHNOLOGY": "Biotechnology", "DEPARTMENT OF BIOTECHNOLOGY": "Biotechnology",
        "DATA SCIENCE": "Data Science", "BCA DS": "BCA Data Science", "CYBER": "Cyber Security", "CYBER SECURITY": "Cyber Security", "DEPARTMENT OF CYBER SECURITY": "Cyber Security",
        "VISCOM": "Visual Communication", "MCA /BCA": "MCA / BCA",
    }
    return aliases.get(raw, clean(value).title())


def rows(ws, start, columns):
    return [r for r in ws.iter_rows(min_row=start, max_col=columns, values_only=True) if r[0] is not None]


def main():
    # The consolidated source is authoritative for year-selected headline counts.
    summary = {
        2024: {"faculty": 172, "publications": 34, "patents": 66, "scholars": 18, "supervisors": 15, "fundedProjects": 5},
        2025: {"faculty": 208, "publications": 59, "patents": 87, "scholars": 52, "supervisors": 33, "fundedProjects": 3},
        2026: {"faculty": 265, "publications": 42, "patents": 66, "scholars": 44, "supervisors": 30, "fundedProjects": 0},
    }

    publications = []
    book = load_workbook(RESEARCH / "FLABS_SCOPUS indexed 2024-2026 (1).xlsx", read_only=True, data_only=True)
    for sheet in book.worksheets:
        year = int(sheet.title)
        for row in rows(sheet, 2, 5):
            publications.append({"id": f"P{year}-{int(row[0])}", "year": year, "authors": clean(row[1]), "title": clean(row[2]), "journal": clean(row[3]), "link": clean(row[4])})

    patents = []
    book = load_workbook(RESEARCH / "Patent 2024-2026 (1).xlsx", read_only=True, data_only=True)
    for sheet in book.worksheets:
        for row in rows(sheet, 5, 6):
            status = clean(row[5]).title() or "Not recorded"
            patents.append({"id": f"T{int(row[4])}-{int(row[0])}", "year": int(row[4]), "department": clean(row[1]), "departmentGroup": department(row[1]), "inventors": clean(row[2]), "title": clean(row[3]), "status": status})
    for year in summary:
        matching = [p for p in patents if p["year"] == year]
        summary[year]["granted"] = sum(p["status"] == "Granted" for p in matching)
        summary[year]["published"] = sum(p["status"] == "Published" for p in matching)

    awards = []
    award_sheet = load_workbook(RESEARCH / "Research Awards 2024-2026 (1).xlsx", read_only=True, data_only=True).active
    for row in rows(award_sheet, 5, 4):
        awards.append({"faculty": clean(row[1]), "award": clean(row[2]), "year": int(row[3])})
    for year in summary:
        summary[year]["awards"] = sum(a["year"] == year for a in awards)

    funded = []
    funded_sheet = load_workbook(YEARWISE / "Funded Project 2024-2026 (Year wise).xlsx", read_only=True, data_only=True).active
    for row in rows(funded_sheet, 4, 7):
        funded.append({"id": int(row[0]), "title": clean(row[1]), "principalInvestigator": clean(row[2]), "coPrincipalInvestigator": clean(row[3]), "agency": clean(row[4]), "amount": clean(row[5]), "year": int(row[6])})

    scholars = []
    scholar_book = load_workbook(YEARWISE / "Research Scholar and Supervisor - Year wise.xlsx", read_only=True, data_only=True)
    for sheet in scholar_book.worksheets:
        year = int(sheet.title)
        for row in rows(sheet, 2, 6):
            scholars.append({"year": year, "supervisor": clean(row[1]), "registration": clean(row[2]), "name": clean(row[3]), "category": clean(row[4]).title(), "department": clean(row[5]), "departmentGroup": department(row[5])})

    data = {"publications": publications, "patents": patents, "awards": awards, "funded": funded, "scholars": scholars, "summary": summary}
    OUT.write_text("window.RESEARCH_DATA = " + json.dumps(data, ensure_ascii=False) + ";\n", encoding="utf-8")
    print(f"Created {OUT.name}: {len(publications)} publications, {len(patents)} patents, {len(awards)} awards, {len(funded)} funded projects, {len(scholars)} scholar records.")


if __name__ == "__main__":
    main()
